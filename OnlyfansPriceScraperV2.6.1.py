from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
import configparser
import time
from selenium.common.exceptions import NoSuchElementException
import os
import argparse
from prompt_toolkit import PromptSession
from prompt_toolkit.completion import WordCompleter
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.cell.cell import Cell

# Maak een ArgumentParser object
parser = argparse.ArgumentParser(description='Select the language for your Python project.')

# Voeg een argument toe voor de taal
parser.add_argument('-l', '--language', type=str, required=False, default='dutch', help='Select the language (for example. "english" or "dutch")')

# Parse de argumenten
args = parser.parse_args()

# Maak een WordCompleter voor de taalargumenten
language_completer = WordCompleter(['english', 'dutch'], ignore_case=True)

# Maak een PromptSession
session = PromptSession()

# Vraag de gebruiker om de taal in te voeren
language = session.prompt('Select language by using tab (default is "dutch"): ', completer=language_completer)
args.language = language if language else 'dutch'

# Taalondersteuning
language_dict = {
 'english': {
   'language_set': 'Language is set on English',
   'no_limited_offer': 'No limited offer found.'
 },
 'dutch': {
   'language_set': 'Taal is ingesteld op Nederlands',
   'no_limited_offer': 'Geen limited offer gevonden.'
 }
}

# Controleer de geselecteerde taal
if args.language in language_dict:
 print(language_dict[args.language]['language_set'])
else:
 print("Unknown language selected. Language is set to Dutch as default.")
 print(language_dict['dutch'])

wb = Workbook()
ws = wb.active

# Voeg gegevens toe aan het werkblad
ws['A1'] = 'Link'
ws['B1'] = 'Price'
ws['C1'] = 'Limited Offer'

def get_onlyfans_prices(profile_url):
  options = Options()
  options.add_argument("--headless")
  driver = webdriver.Firefox(options=options)

  driver.get(profile_url)

  time.sleep(5)

  driver.implicitly_wait(10)

  price_element = None
  limited_offer_element = None

  try:
      price_element = driver.find_element(By.CLASS_NAME, 'b-wrap-btn-text')
  except NoSuchElementException:
      try:
          price_element = driver.find_element(By.CLASS_NAME, 'b-offer-join')
      except NoSuchElementException:
          print(f"Geen van de elementen gevonden op {profile_url}")
          return profile_url, 'Prijs niet gevonden op de opgegeven URL.', 'Geen limited offer opgegeven.'

  try:
      limited_offer_element = driver.find_element(By.CLASS_NAME, 'b-offer-join__content')
  except NoSuchElementException:
      pass

  if price_element:
      price = price_element.text.strip()
  else:
      price = 'Prijs niet gevonden op de opgegeven URL.'

  if limited_offer_element:
      limited_offer = limited_offer_element.text.strip()
  else:
      limited_offer = language_dict[args.language]['no_limited_offer']

  driver.quit()

  return profile_url, price, limited_offer


config = configparser.ConfigParser()
config.read('Settings.ini')

user_agent = config.get('Settings', 'user_agent')

options = Options()
options.add_argument(f"user-agent={user_agent}")

config.read('Linkconfig.ini')

urls = [config.get('URLs', url) for url in config['URLs'] if url != 'user_agent']

output_dir = "output"
if not os.path.exists(output_dir):
 try:
   os.mkdir(output_dir)
   print(f"Successfully created the directory {output_dir}")
 except OSError:
   print(f"Creation of the directory {output_dir} failed")
else:
 print(f"Directory {output_dir} already exists")

output_file = os.path.join(output_dir, "output.xlsx")

if os.path.isfile(output_file):
 os.remove(output_file)
 print(f"Old data in {output_file} removed")
else:
 print(f"No old data to remove in {output_file}")

# Filter alleen de geldige URL's
valid_urls = [url for url in urls if url and urlparse(url).netloc]

for url in valid_urls:
 profile_url, price, limited_offer = get_onlyfans_prices(url)
 price = price.replace('\n', ' ')
 limited_offer = limited_offer.replace('\n', ' ')
 if args.language == 'english':
   print('Price found at {}: {}'.format(url, price))
   print('Limited Offer found at {}: {}'.format(url, limited_offer))
 else:
   if args.language == 'dutch':
       print('Prijs gevonden op {}: {}'.format(url, price))
       print('Limited Offer gevonden op {}: {}'.format(url, limited_offer))
 hyperlink = Hyperlink(display=profile_url, ref=profile_url)
 ws.append([profile_url, price, limited_offer])
 ws.cell(row=ws.max_row, column=1).hyperlink = hyperlink
 ws.cell(row=ws.max_row, column=1).style = 'Hyperlink'

wb.save(filename=output_file)
