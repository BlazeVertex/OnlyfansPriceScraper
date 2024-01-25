from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
import configparser
import time
from selenium.common.exceptions import NoSuchElementException
import os
import argparse
from prompt_toolkit import PromptSession
from prompt_toolkit.completion import WordCompleter
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
import json
import re
import pandas as pd
import warnings
warnings.filterwarnings("error", category=FutureWarning)

# Maak een ArgumentParser object
parser = argparse.ArgumentParser(description='Select the language for your Python project.')

# Voeg een argument toe voor de taal
parser.add_argument('-l', '--language', type=str, required=False, default='dutch', help='Select the language (for example. "english" or "dutch" , "french" , "german")')

# Parse de argumenten
args = parser.parse_args()

# Maak een WordCompleter voor de taalargumenten
language_completer = WordCompleter(['english', 'dutch', 'french', 'german', 'spanish'], ignore_case=True)

# Maak een PromptSession
session = PromptSession()

# Vraag de gebruiker om de taal in te voeren
language = session.prompt('Select language by using tab (default is "dutch"): ', completer=language_completer)
args.language = language if language else 'dutch'

# Lees de taalinstellingen van het JSON bestand
with open('Language.json') as f:
  language_dict = json.load(f)

# Controleer de geselecteerde taal
if args.language in language_dict:
 print(language_dict[args.language]['language_set'])
else:
 print("Unknown language selected. Language is set to Dutch as default.")
 print(language_dict['dutch'])

wb = Workbook()
ws = wb.active

# Voeg gegevens toe aan het werkblad
ws['A1'] = 'Link'
ws['B1'] = 'Price'
ws['C1'] = 'Limited Offer'

def get_onlyfans_prices(profile_url):
  options = Options()
  options.add_argument("--headless")
  driver = webdriver.Firefox(options=options)

  driver.get(profile_url)

  time.sleep(5)

  driver.implicitly_wait(10)

  price_element = None
  limited_offer_element = None

  try:
      price_element = driver.find_element(By.CLASS_NAME, 'b-wrap-btn-text')
  except NoSuchElementException:
      try:
          price_element = driver.find_element(By.CLASS_NAME, 'b-offer-join')
      except NoSuchElementException:
          print(f"Geen van de elementen gevonden op {profile_url}")
          return profile_url, 'Prijs niet gevonden op de opgegeven URL.', 'Geen limited offer opgegeven.'

  try:
      limited_offer_element = driver.find_element(By.CLASS_NAME, 'b-offer-join__content')
  except NoSuchElementException:
      pass

  if price_element:
      price = price_element.text.strip()
  else:
      price = 'Prijs niet gevonden op de opgegeven URL.'

  if limited_offer_element:
      limited_offer = limited_offer_element.text.strip()
  else:
      limited_offer = language_dict[args.language]['no_limited_offer']

  driver.quit()

  return profile_url, price, limited_offer

def extract_price(s):
 match = re.search(r'\$([0-9.,]+)', s)
 if match:
     return float(match.group(1).replace(',', ''))
 else:
     return None

config = configparser.ConfigParser()
config.read('Settings.ini')

user_agent = config.get('Settings', 'user_agent')
high_price_first = config.getboolean('Settings', 'high_price_first')
low_price_first = config.getboolean('Settings', 'low_price_first')

options = Options()
options.add_argument(f"user-agent={user_agent}")

config.read('Linkconfig.ini')

urls = [config.get('URLs', url) for url in config['URLs'] if url != 'user_agent']

output_dir = "output"
if not os.path.exists(output_dir):
 try:
   os.mkdir(output_dir)
   print(f"Successfully created the directory {output_dir}")
 except OSError:
   print(f"Creation of the directory {output_dir} failed")
else:
 print(f"Directory {output_dir} already exists")

output_file = os.path.join(output_dir, "temp.xlsx")

if os.path.isfile(output_file):
 os.remove(output_file)
 print(f"Old data in {output_file} removed")
else:
 print(f"No old data to remove in {output_file}")

# Filter alleen de geldige URL's
valid_urls = [url for url in urls if url and urlparse(url).netloc]

df = pd.DataFrame(columns=['Link', 'Price', 'Limited Offer'])

for url in valid_urls:
  profile_url, price, limited_offer = get_onlyfans_prices(url)
  price = price.replace('\n', ' ')
  limited_offer = limited_offer.replace('\n', ' ')
  
  # Extraheer de prijs uit de 'Price' kolom
  price = extract_price(price)
  
  if profile_url and price is not None and limited_offer:
      # Voeg de gegevens toe aan de DataFrame
      new_row = pd.Series({'Link': profile_url, 'Price': price, 'Limited Offer': limited_offer})
      df = pd.concat([df, new_row.to_frame().transpose()], ignore_index=True)
      
      if args.language == 'english':
          print(f'{language_dict[args.language]["price_found"]} {url}: {price}')
          print(f'Limited Offer found at {url}: {limited_offer}')
      elif args.language == 'dutch':
          print(f'{language_dict[args.language]["price_found"]} {url}: {price}')
          print(f'Limited Offer gevonden op {url}: {limited_offer}')
      elif args.language == 'french':
          print(f'{language_dict[args.language]["price_found"]} {url}: {price}')
          print(f'Offre limitée trouvée à {url}: {limited_offer}')
      elif args.language == 'german':
          print(f'{language_dict[args.language]["price_found"]} {url}: {price}')
          print(f'Begrenztes Angebot gefunden bei {url}: {limited_offer}')
      elif args.language == 'spanish':
          print(f'{language_dict[args.language]["price_found"]} {url}: {price}')
          print(f'Oferta limitada encontrada {url}: {limited_offer}')

# Sorteer en sla de DataFrame op
if high_price_first:
   df = df.sort_values(by='Price', ascending=False)
elif low_price_first:
   df = df.sort_values(by='Price', ascending=True)
sorted_output_file = os.path.join(output_dir, 'sorted_output.xlsx')
df.to_excel(sorted_output_file, index=False)

# Laad het Excel-bestand
wb = load_workbook(sorted_output_file)
ws = wb.active

# Stel de breedte van de kolommen in
for col in ws.columns:
   max_length = 0
   column = col[0].column_letter
   for cell in col:
       try:
           if len(str(cell.value)) > max_length:
               max_length = len(cell.value)
       except:
           pass
   adjusted_width = (max_length + 2)
   ws.column_dimensions[column].width = adjusted_width

# Sla het Excel-bestand op
wb.save(filename=sorted_output_file)
